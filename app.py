import streamlit as st
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class InventoryTracker:
    def __init__(self):
        self.roll_types = [
            "BOPP Transparent 30u720",
            "BOPP Métallisé",
            "CPP Standard",
            "PE Alimentaire",
            "Polyéthylène",
            "Polypropylène",
            "Film Alimentaire",
            "Film Industriel",
            "Autre"
        ]

        self.product_names = [
            "Rouleau Standard",
            "Rouleau Transparent",
            "Rouleau Métallisé",
            "Rouleau Alimentaire",
            "Rouleau Spécial",
            "Film Technique",
            "Film Protection",
            "Autre"
        ]

        self.companies = [
            "Société A",
            "Société B",
            "Société C",
            "Nouveau Client",
            "Client Existant"
        ]

        current_year = datetime.now().year % 100
        self.delivery_numbers = [f"{num:02d}/{current_year}" for num in range(1, 100)]

        if 'entries' not in st.session_state:
            st.session_state.entries = []

    def calculate_tare(self, num_rolls, width_mm):
        tare = 3.22 / 2000 * num_rolls * width_mm
        return round(tare, 2)

    def add_entry(self, date, delivery_number, company, phone, num_rolls, roll_type, product_name, width_mm, gross_weight):
        tare = self.calculate_tare(num_rolls, width_mm)
        net_weight = round(gross_weight - tare, 2)

        entry = {
            'Date': date,
            'Numéro de Livraison': delivery_number,
            'Nom de l\'Entreprise': company,
            'Téléphone': phone,
            'Nombre de Rouleaux': num_rolls,
            'Type de Rouleau': roll_type,
            'Nom du Produit': product_name,
            'Largeur (mm)': width_mm,
            'Poids Brut (kg)': gross_weight,
            'Tare (kg)': tare,
            'Poids Net (kg)': net_weight
        }

        st.session_state.entries.append(entry)
        st.success("Entrée ajoutée avec succès!")

    def export_to_excel(self, df):
        output_path = f"Inventaire_{datetime.now().strftime('%Y%m%d')}.xlsx"
        df.to_excel(output_path, index=False, sheet_name='Inventaire')
        workbook = openpyxl.load_workbook(output_path)
        worksheet = workbook['Inventaire']

        header_font = Font(bold=True)
        header_alignment = Alignment(wrap_text=True, vertical='top')
        header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
        header_border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))

        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border

        # Add a footer note in Excel
        worksheet['A' + str(len(df) + 3)] = "Développé par Mohammed EL Hassani"
        worksheet['A' + str(len(df) + 3)].font = Font(italic=True, color="0000FF")

        workbook.save(output_path)
        return output_path

    def display_interface(self):
        st.title("📦 Suivi des Stocks de Rouleaux")
        st.caption("Développé par Mohammed EL Hassani")  # Display at the top

        with st.form("inventory_form"):
            col1, col2 = st.columns(2)

            with col1:
                date = st.date_input("Date")

                # Delivery number selection
                delivery_number = st.selectbox(
                    "Numéro de Livraison",
                    options=self.delivery_numbers + ['Autre']
                )
                if delivery_number == 'Autre':
                    delivery_number = st.text_input("Saisissez le Numéro de Livraison")

                # Allow custom company names
                company = st.text_input("Nom de l'Entreprise (Choisissez ou saisissez)")
                if company.strip() == "":
                    company = st.selectbox(
                        "Sélectionnez une Entreprise",
                        options=self.companies
                    )

                phone = st.text_input("Téléphone")

            with col2:
                num_rolls = st.number_input("Nombre de Rouleaux", min_value=1, value=1)

                # Allow custom roll types
                roll_type = st.text_input("Type de Rouleau (Choisissez ou saisissez)")
                if roll_type.strip() == "":
                    roll_type = st.selectbox(
                        "Sélectionnez un Type de Rouleau",
                        options=self.roll_types
                    )

                # Allow custom product names
                product_name = st.text_input("Nom du Produit (Choisissez ou saisissez)")
                if product_name.strip() == "":
                    product_name = st.selectbox(
                        "Sélectionnez un Produit",
                        options=self.product_names
                    )

                width_mm = st.number_input("Largeur (mm)", min_value=1)
                gross_weight = st.number_input("Poids Brut (kg)", min_value=0.0, step=0.1)

            # Submit button
            submit = st.form_submit_button("Ajouter l'Entrée")
            if submit:
                self.add_entry(date, delivery_number, company, phone, num_rolls, roll_type, product_name, width_mm, gross_weight)

        # Display inventory entries
        if st.session_state.entries:
            df = pd.DataFrame(st.session_state.entries)

            # Format the date column
            df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%d/%m/%Y')
            st.subheader("📋 Entrées Actuelles")
            st.dataframe(df, use_container_width=True)

            col1, col2 = st.columns(2)

            with col1:
                csv = df.to_csv(index=False).encode('utf-8')
                st.download_button("📥 Télécharger CSV", data=csv, file_name="inventory.csv", mime="text/csv")

            with col2:
                excel_path = self.export_to_excel(df)
                with open(excel_path, "rb") as file:
                    st.download_button(
                        "📥 Télécharger Excel",
                        data=file,
                        file_name=excel_path,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

        # Footer note in app
        st.markdown("<hr>", unsafe_allow_html=True)
        st.caption("© 2025 Développé par Mohammed EL Hassani")

def main():
    tracker = InventoryTracker()
    tracker.display_interface()

if __name__ == "__main__":
    main()
